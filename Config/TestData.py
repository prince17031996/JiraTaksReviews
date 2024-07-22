import openpyxl

class TestData():
    def Links(self):
        workbook = openpyxl.load_workbook("C:\\Users\\PRaj7\\PycharmProjects\\DORUserStory\\Config\\DOR_Check.xlsx")
        sheet = workbook["URL"]
        total_rows = sheet.max_row
        total_colmns = sheet.max_column
        l = []
        for r in range(2, total_rows + 1):
            for c in range(1, total_colmns + 1):
                l.append(sheet.cell(r, c).value)
        return l

obj=TestData()
print("User Story",obj.Links())






class TestScriptData:
    def Links(self):
        workbook = openpyxl.load_workbook("C:\\Users\\PRaj7\\PycharmProjects\\DORUserStory\\Config\\script_Check.xlsx")
        sheet = workbook["URL"]
        total_rows = sheet.max_row
        total_colmns = sheet.max_column
        l = []
        for r in range(2, total_rows + 1):
            for c in range(1, total_colmns + 1):
                l.append(sheet.cell(r, c).value)
        return l


obj=TestScriptData()
print("Test Script",obj.Links())

class TestDataBug:
    def Links(self):
        workbook = openpyxl.load_workbook("C:\\Users\\PRaj7\\PycharmProjects\\DORUserStory\\Config\\Bug_Check.xlsx")
        sheet = workbook["URL"]
        total_rows = sheet.max_row
        total_colmns = sheet.max_column
        l = []
        for r in range(2, total_rows + 1):
            for c in range(1, total_colmns + 1):
                l.append(sheet.cell(r, c).value)
        return l

obj=TestDataBug()
print("Bug Data",obj.Links())

class TestExecutionData:
    def Links(self):
        workbook = openpyxl.load_workbook("C:\\Users\\PRaj7\\PycharmProjects\\DORUserStory\\Config\\scriptExecutionIssueKey_check.xlsx")
        sheet = workbook["URL"]
        total_rows = sheet.max_row
        total_colmns = sheet.max_column
        l = []
        for r in range(2, total_rows + 1):
            for c in range(1, total_colmns + 1):
                l.append(sheet.cell(r, c).value)
        return l


obj=TestExecutionData()
print("Test Execution Data",obj.Links())
#f"https://jira.jnj.com/secure/XrayExecuteTest!default.jspa?testExecIssueKey={executionKey}&testIssueKey={testIssueKey}"

class TestExecutionTestIssueKey:
    def Links(self):
        workbook = openpyxl.load_workbook("C:\\Users\\PRaj7\\PycharmProjects\\DORUserStory\\Config\\scriptPostExecutionTestIssueKey_check.xlsx")
        sheet = workbook["URL"]
        total_rows = sheet.max_row
        total_colmns = sheet.max_column
        l = []

        linkText=["https://jira.jnj.com/secure/XrayExecuteTest!default.jspa?testExecIssueKey=","&testIssueKey="]
        for r in range(2, total_rows + 1):
            str = ""
            for c in range(1, total_colmns + 1):
                str+=(f"{linkText[c-1]}{sheet.cell(r, c).value}")
            l.append(str)
        return l

obj=TestExecutionTestIssueKey()
print("TestIssueKey",obj.Links())

