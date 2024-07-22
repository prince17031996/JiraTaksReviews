import time

import openpyxl
from openpyxl.styles import PatternFill
from selenium.common import NoSuchElementException, TimeoutException
from selenium.webdriver.common.by import By

from Config.TestData import TestData
from Config.locators import Locators
from Pages.basePage import Basepage


class TESTSCRIPTEEXECUTIONTESTISSUE(Basepage):

    """"def __init__(self, driver):
        super().__init__(driver)
        obj = TestData()
        self.urls = obj.Links()"""""

    def run_tests(self, url):
        self.driver.get(url)

        # Load the workbook and select the "Result" worksheet or create it if it doesn't exist
        workbook = openpyxl.load_workbook(r'C:\Users\PRaj7\PycharmProjects\DORUserStory\Config\scriptPostExecutionTestIssueKey_check.xlsx')
        if 'Result' not in workbook.sheetnames:
            workbook.create_sheet(title='Result')
        worksheet = workbook['Result']

        # Write headers to the first row of the "Result" worksheet if it's a new workbook
        if worksheet.max_row == 1:
            headers = ["Link", "Actual Result", "Evidence", "Step State"]
            for col, header in enumerate(headers, start=1):
                worksheet.cell(row=1, column=col, value=header)

        # Write the URL to the "Result" worksheet
        row = worksheet.max_row + 1
        worksheet.cell(row=row, column=1, value=url)

        # Perform tests and write results to the "Result" worksheet
        test_results = self.get_test_results()
        for col, result in enumerate(test_results.values(), start=2):
            cell = worksheet.cell(row=row, column=col)
            cell.value = result
            if result == "Passed":
                cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

            else:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Save the workbook
        workbook.save(r'C:\Users\PRaj7\PycharmProjects\DORUserStory\Config\scriptPostExecutionTestIssueKey_check.xlsx')

        # Return the test results
        return test_results

    def ActualResultField(self):
        result=""
        for x in range(1, 100):
            time.sleep(1)
            try:
                scroll=self.driver.find_element(By.XPATH, f"(//div[@class='raven-test-step-container'])[{x}]")
                #clicking on actual result button
                button = self.driver.find_element(By.XPATH, f"(//button[contains(text(),'Actual Result')])[{x}]")


                button.click()
                time.sleep(1)
                # Construct XPath for each actual result element
                actualResult =self.driver.find_element(By.XPATH,
                                                   f"(//div[@class='raven-field-wiki-view']//div[@class='field-content text-wrap'])[{x}]")
                self.driver.execute_script("arguments[0].scrollIntoView();", scroll)
                time.sleep(1)
                if len(actualResult.text)>5 :
                    pass
                else:
                    result+=f"{x}"
                    result+= ","


            except (NoSuchElementException, TimeoutException):
                print(f"Timeout waiting for actualResult {x} to be visible")
                break

        if len(result)>0:
            return f"Update for Step {result}"
        else:
            return "Passed"

    def EvidenceField(self):
        result = ""
        for x in range(1, 100):
            time.sleep(1)
            try:

                # Construct XPath for each Evidence
                evidence = self.driver.find_element(By.XPATH,
                                                        f"(//div[@id='evidences-actions']//span[@class='count'])[{x}]")
                self.driver.execute_script("arguments[0].scrollIntoView();", evidence)

                if (evidence.text!="(0)"):
                    pass
                else:
                    result += f"{x}"
                    result += ","


            except (NoSuchElementException, TimeoutException):
                print(f"Timeout waiting for evidence {x} to be visible")
                break

        if len(result) > 0:
            return f"Update for Step {result}"
        else:
            return "Passed"

    def StepStateField(self):
        result = ""
        for x in range(1, 100):
            time.sleep(1)
            try:
                scroll = self.driver.find_element(By.XPATH, f"(//div[@class='raven-test-step-container'])[{x}]")

                # Construct XPath for each actual result element
                stepState = self.driver.find_element(By.XPATH,
                                                        f"(//span[@class='raven-status-name status-tipsy'])[{x+1}]")
                self.driver.execute_script("arguments[0].scrollIntoView();", scroll)
                print("stepState result is ", stepState.text)
                time.sleep(1)
                if (stepState.text=="PASS") or (stepState.text=="NA"):
                    pass
                else:
                    result += f"{x}"
                    result += ","


            except (NoSuchElementException, TimeoutException):
                print(f"Timeout waiting for actualResult {x} to be visible")
                break
        if len(result)>0:
            return f"Update for Step {result}"
        else:
            return "Passed"






    def get_test_results(self):
        actualResult_text = self.ActualResultField()
        evidence_text = self.EvidenceField()
        stepState_text = self.StepStateField()


        return {
            "Actual Result": actualResult_text,
            "Evidence": evidence_text,
            "Step State": stepState_text,


        }
