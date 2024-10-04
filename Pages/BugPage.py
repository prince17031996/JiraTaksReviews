import openpyxl
from openpyxl.styles import PatternFill

from Config.TestData import TestData
from Config.locators import Locators
from Pages.basePage import Basepage


class BUG(Basepage):

    """"def __init__(self, driver):
        super().__init__(driver)
        obj = TestData()
        self.urls = obj.Links()"""""

    def run_tests(self, url):
        self.driver.get(url)
        self.driver.maximize_window()

        # Load the workbook and select the "Result" worksheet or create it if it doesn't exist
        workbook = openpyxl.load_workbook(r'C:\Users\PRaj7\PycharmProjects\DORUserStory\Config\Bug_Check.xlsx')
        if 'Result' not in workbook.sheetnames:
            workbook.create_sheet(title='Result')
        worksheet = workbook['Result']

        # Write headers to the first row of the "Result" worksheet if it's a new workbook
        if worksheet.max_row == 1:
            headers = ["URL", "Title", "Type", "Affected Version", "Fix Version", "Story Point",
                       "Description","Priority","Assignee","Severity","Root Cause"]
            for col, header in enumerate(headers, start=1):
                worksheet.cell(row=1, column=col, value=header)

        # Write the URL to the "Result" worksheet
        row = worksheet.max_row + 1
        worksheet.cell(row=row, column=1, value=url)

        # Perform tests and write results to the "Result" worksheet
        test_results = self.get_test_results()
        for col, result in enumerate(test_results.values(), start=2):
            cell = worksheet.cell(row=row, column=col)
            cell.value = result
            if result == "Passed":
                cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

            else:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Save the workbook
        workbook.save(r'C:\Users\PRaj7\PycharmProjects\DORUserStory\Config\Bug_Check.xlsx')

        # Return the test results
        return test_results

    def title(self):
        if self.isVisible(Locators.typeButton):
            return "Passed"
        else:
            return "Failed"


    def typeField(self):
        if self.getText(Locators.typeText)=="Bug":
            return "Passed"
        else:
            return "Update the type as Story"

    def affectedVersonField(self):
        if self.getText(Locators.affectedVersionText)=="None":
            return "Update the Affected version"
        else:
            return "Passed"

    def fixVersionField(self):
        if self.getText(Locators.fixVersionText)=="None":
            return "Update the fix version"
        else:
            return "Passed"


    def storyPointField(self):
        if self.isVisible(Locators.storyPointButton):
            return "Passed"
        else:
            return "Add the Story Point"


    def descriptionField(self):
        if len(self.getText(Locators.descriptionText))>30:
            return "Passed"
        else:
            return "Add description"
    def priorityField(self):
        if self.isVisible(Locators.priorityButton):
            return "Passed"
        else:
            return "Add the Priority"


    def assigneeField(self):
        if self.getText(Locators.assigneeText)=="Unassigned":
            return "Update the assignee Name"
        else:
            return "Passed"


    def severityField(self):
        if self.isVisible(Locators.severityButton):
            return "Passed"
        else:
            return "Add severity"

    def rootCauseField(self):
        if self.isVisible(Locators.rootCauseButton):
            return "Passed"
        else:
            return "Add rootcause"









    def get_test_results(self):
        title = self.title()
        type_text = self.typeField()
        affected_version_text = self.affectedVersonField()
        fix_version_text = self.fixVersionField()
        story_point_text = self.storyPointField()
        description_text = self.descriptionField()
        priority_text = self.priorityField()
        assignee_text = self.assigneeField()
        severity_text=self.severityField()
        root_cause_text=self.rootCauseField()

        #due_date_text=self.dueDateField()
        #compliance_type=self.complianceField()

        return {
            "Title": title,
            "Type": type_text,
            "Affected Version": affected_version_text,
            "Fix Version": fix_version_text,
            "Story Point": story_point_text,
            "Description": description_text,
            "Priority": priority_text,
            "Assginne":assignee_text,
            "Severity":severity_text,
            "Root Cause":root_cause_text
           # "Due Date":due_date_text,
            #"ComplianceType":compliance_type

        }
