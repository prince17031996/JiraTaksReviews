import time

import openpyxl
from openpyxl.styles import PatternFill
from selenium.common import NoSuchElementException, TimeoutException
from selenium.webdriver.common.by import By

from Config.TestData import TestData
from Config.locators import Locators
from Pages.basePage import Basepage


class TESTSCRIPTEEXECUTION(Basepage):

    """"def __init__(self, driver):
        super().__init__(driver)
        obj = TestData()
        self.urls = obj.Links()"""""

    def run_tests(self, url):
        self.driver.get(url)

        # Load the workbook and select the "Result" worksheet or create it if it doesn't exist
        workbook = openpyxl.load_workbook(r'C:\Users\PRaj7\PycharmProjects\DORUserStory\Config\scriptExecutionIssueKey_check.xlsx')
        if 'Result' not in workbook.sheetnames:
            workbook.create_sheet(title='Result')
        worksheet = workbook['Result']

        # Write headers to the first row of the "Result" worksheet if it's a new workbook
        if worksheet.max_row == 1:
            headers = ["URL", "Title", "Type", "Affected Version", "Fix Version", "Description"
                , "Priority", "Approval Workflow", "Assignee/Reporter", "Sprint", "Issue Resolution",
                       "Test Category", "Test Type", "Issue Link", "Execution Link"]
            for col, header in enumerate(headers, start=1):
                worksheet.cell(row=1, column=col, value=header)

        # Write the URL to the "Result" worksheet
        row = worksheet.max_row + 1
        worksheet.cell(row=row, column=1, value=url)

        # Perform tests and write results to the "Result" worksheet
        test_results = self.get_test_results()
        for col, result in enumerate(test_results.values(), start=2):
            cell = worksheet.cell(row=row, column=col)
            cell.value = result
            if result == "Passed":
                cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

            else:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Save the workbook
        workbook.save(r'C:\Users\PRaj7\PycharmProjects\DORUserStory\Config\scriptExecutionIssueKey_check.xlsx')

        # Return the test results
        return test_results

    def title(self):
        if self.isVisible(Locators.typeButton):
            return "Passed"
        else:
            return "Failed"


    def typeField(self):
        if self.getText(Locators.typeText)=="Test":
            return "Passed"
        else:
            return "Update the type as Test"

    def affectedVersonField(self):
        if self.getText(Locators.affectedVersionText)=="None":
            return "Update the Affected version"
        else:
            return "Passed"

    def fixVersionField(self):
        if self.getText(Locators.fixVersionText)=="None":
            return "Update the fix version"
        else:
            return "Passed"


    def descriptionField(self):
        if self.isVisible(Locators.descriptionButton):
            return "Passed"
        else:
            return "Add the description"
    def priorityField(self):
        if self.isVisible(Locators.priorityButton):
            return "Passed"
        else:
            return "Add the Priority"

    def approvalField(self):
        if self.getText(Locators.approvalText)=="NONE":
            return "Update the approval Workflow"
        else:
            return "Passed"

    def assigneeReporterField(self):
        if self.getText(Locators.assigneeText)=="Unassigned":
            return "Update the assignee Name"
        else:
            return "Passed"

    def sprintField(self):
        if self.isVisible(Locators.sprintButton):
            return "Passed"
        else:
            return "Update Sprint Details"


    def issueResolutionField(self):
        if self.isVisible(Locators.issueResolutionButton):
            return "Passed"
        else:
            return "Add Issue Resolution"

    def categoryTestField(self):
        if self.isVisible(Locators.testCategoryButton):
            return "Passed"
        else:
            return "Add test Category"

    def typeTestField(self):
        type=''
        if self.isVisible(Locators.testTypeButton):
            return "Passed"
        else:
            return "Update Test Type"

    def issueLinkField(self):
        if self.isVisible(Locators.issueLinkButton):
            return "Passed"
        else:
            return "Update the Issue Link"

    def executionField(self):
        if self.isVisible(Locators.testExecutionButton):
            return "Passed"
        else:
            return "Test Execution Not Created"


    def actualResultField(self):

       # actualResult=(By.XPATH,f"(//div[@class='raven-field-wiki-view']//child::div[@class='field-content text-wrap'])[{x}]//p")
       for x in range(1, 100):
            try:
                actualResult=(By.XPATH,f"(//div[@class='raven-field-wiki-view']//child::div[@class='field-content text-wrap'])[{x}]//p")
                if self.isVisible(actualResult):
                    print(self.getText(actualResult))

                self.driver.execute_script("arguments[0].scrollIntoView();", actualResult)

            except(NoSuchElementException,TimeoutException)as e:
                print("error")
                break




    def get_test_results(self):
        """title = self.title()
        type_text = self.typeField()
        affected_version_text = self.affectedVersonField()
        fix_version_text = self.fixVersionField()
        description_text = self.descriptionField()
        priority_text = self.priorityField()
        approval_text = self.approvalField()
        assigneeReporter_text = self.assigneeReporterField()
        sprint_text = self.sprintField()
        issueResolution_text = self.issueResolutionField()
        testCategory_text = self.categoryTestField()
        testType_text = self.typeTestField()
        issueLink_text = self.issueLinkField()
        execution_text = self.executionField()"""

        #action_text = self.actionField()
        #expectedResult_text= self.expectedResultField()

        return {
            """"Title": title,
            "Type": type_text,
            "Affected Version": affected_version_text,
            "Fix Version": fix_version_text,
            "Description": description_text,
            "Priority": priority_text,
            "Approval" : approval_text,
            "AssginneReporter":assigneeReporter_text,
            "Sprint" : sprint_text,
            "IssueResolution": issueResolution_text,
            "TestCategory": testCategory_text,
            "TestType": testType_text,
            "IssueLink":issueLink_text,
            "ExecutionLink":execution_text,"""

        }
